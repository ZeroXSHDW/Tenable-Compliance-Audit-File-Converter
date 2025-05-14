import os
import sys
import json
import logging
import re
import shutil
import hashlib
from typing import List, Set
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
import chardet
from tqdm import tqdm

__version__ = "2.0.0"
__changelog__ = """
2.0.0:
- Initial implementation for audit file processing utilities
- Added openpyxl support for XLSX output
- Integrated logging and config management
- Updated FILENAME_PATTERN to preserve version numbers
- Added sanitization for XLSX cell values
- Added temporary config file handling
- Disabled config tag updates to protect config.json
- Added config checksum verification
- Enhanced sanitization logging to track field and file
- Elevated config checksum and sanitization logs to INFO level
"""

FILENAME_PATTERN = re.compile(r'[^a-zA-Z0-9_.-]+')

def setup_logger(name: str, log_file: str, config: dict) -> logging.Logger:
    """Set up a logger with file and console output."""
    logger = logging.getLogger(name)
    logger.setLevel(getattr(logging, config.get('logging', {}).get('level', 'INFO')))
    
    if not logger.handlers:
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setFormatter(logging.Formatter(config.get('logging', {}).get('format', '%(asctime)s - %(levelname)s - %(message)s')))
        logger.addHandler(file_handler)
        
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setFormatter(logging.Formatter('%(levelname)s: %(message)s'))
        logger.addHandler(console_handler)
    
    return logger

def get_config_checksum(config_path: str) -> str:
    """Calculate SHA256 checksum of config file."""
    try:
        with open(config_path, 'rb') as f:
            return hashlib.sha256(f.read()).hexdigest()
    except Exception as e:
        logging.getLogger('audit_utils').error(f"Failed to calculate checksum for {config_path}: {str(e)}")
        return ""

def cached_config(config_path: str = "config.json") -> dict:
    """Load and cache the configuration file, using a temporary copy."""
    temp_config_path = "config_temp.json"
    logger = logging.getLogger('audit_utils')
    original_checksum = get_config_checksum(config_path)
    logger.info(f"Original config checksum: {original_checksum}")
    
    try:
        shutil.copyfile(config_path, temp_config_path)
        logger.info(f"Created temporary config file: {temp_config_path}")
        with open(temp_config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        return config
    except Exception as e:
        raise RuntimeError(f"Failed to load config {config_path}: {str(e)}")

def verify_config_unchanged(config_path: str, original_checksum: str) -> None:
    """Verify config file was not modified."""
    logger = logging.getLogger('audit_utils')
    current_checksum = get_config_checksum(config_path)
    if current_checksum != original_checksum:
        logger.warning(f"Config file {config_path} modified! Original checksum: {original_checksum}, Current: {current_checksum}")
    else:
        logger.info(f"Config file {config_path} unchanged, checksum: {current_checksum}")

def detect_encoding(file_path: str, logger: logging.Logger) -> tuple[str, str]:
    """Detect file encoding and read content."""
    try:
        with open(file_path, 'rb') as f:
            raw_data = f.read()
        result = chardet.detect(raw_data)
        encoding = result['encoding'] or 'utf-8'
        try:
            content = raw_data.decode(encoding)
        except UnicodeDecodeError:
            logger.warning(f"Failed to decode {file_path} with {encoding}, falling back to utf-8 with errors ignored")
            content = raw_data.decode('utf-8', errors='ignore')
        return encoding, content
    except Exception as e:
        logger.error(f"Failed to detect encoding for {file_path}: {str(e)}")
        return 'utf-8', ""

def validate_json_file(json_file: str, logger: logging.Logger) -> bool:
    """Validate JSON file integrity."""
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            json.load(f)
        return True
    except (json.JSONDecodeError, IOError) as e:
        logger.error(f"JSON validation failed for {json_file}: {str(e)}")
        return False

def update_config_tags(tags: Set[str], config: dict, logger: logging.Logger) -> None:
    """Disabled: No updates to config to protect original config.json."""
    logger.debug(f"Config tag updates disabled to protect config.json. Ignored tags: {tags}")

def ensure_dependencies(logger: logging.Logger, output_format: str, verbose: bool) -> None:
    """Ensure required dependencies are installed."""
    if output_format == 'xlsx' and not OPENPYXL_AVAILABLE:
        logger.error("openpyxl is not installed. Cannot generate XLSX output.")
        raise ImportError("openpyxl is not installed")
    if verbose:
        logger.info(f"Dependencies checked: openpyxl={'available' if OPENPYXL_AVAILABLE else 'missing'}, chardet={'available' if 'chardet' in sys.modules else 'missing'}, tqdm={'available' if 'tqdm' in sys.modules else 'missing'}")

def sanitize_cell_value(value: str, field: str, file_path: str) -> str:
    """Sanitize cell value by removing invalid characters for XLSX."""
    logger = logging.getLogger('audit_utils')
    if not isinstance(value, str):
        value = str(value)
    sanitized = re.sub(r'[^\x20-\x7E\xA0-\xFFFF]', '[INVALID_CHAR]', value)
    if sanitized != value:
        logger.info(f"Sanitized cell value in {file_path}, field '{field}': {value[:50]}... to {sanitized[:50]}...")
    else:
        logger.info(f"No sanitization needed for {file_path}, field '{field}'")
    return sanitized

def write_xlsx(file_path: str, fields: List[str], items: List[dict], config: dict, logger: logging.Logger) -> None:
    """Write items to an XLSX file with styling and sanitization."""
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl is not installed, cannot write XLSX")
        raise ImportError("openpyxl is not installed")
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit Data"
        
        for col, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = field
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        for row, item in enumerate(tqdm(items, desc="Writing XLSX", disable=not config.get('output', {}).get('parallel_processing', True)), 2):
            for col, field in enumerate(fields, 1):
                cell = ws.cell(row=row, column=col)
                value = item.get(field, '')
                cell.value = sanitize_cell_value(value, field, file_path)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)
            ws.column_dimensions[column].width = adjusted_width
        
        for row in ws.rows:
            ws.row_dimensions[row[0].row].height = 15
        
        wb.save(file_path)
        logger.debug(f"XLSX file saved: {file_path}")
    except Exception as e:
        logger.error(f"Failed to write XLSX {file_path}: {str(e)}")
        raise