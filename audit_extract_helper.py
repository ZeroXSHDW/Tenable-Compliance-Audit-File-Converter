import os
import sys
import json
import logging
import time
import argparse
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import Dict, List
from html import escape

__version__ = "2.0.5"
__changelog__ = """
2.0.0:
- Initial implementation for converting JSON to XLSX
- Integrated with audit_utils for XLSX writing and dependency checks
- Enforces XLSX output, removes CSV support
- Sanitizes cell values for XLSX compatibility
- Logs processing steps and sanitization
2.0.1:
- Fixed 'format' error by using output_format
- Updated to support new folder structure
2.0.2:
- Added XLSX conversion logging with timing to debug/xlsx_conversion_log.txt
2.0.3:
- Added HTML output generation
- Uses platform-specific output directories
2.0.4:
- Added enhanced XLSX formatting (bold headers, auto-width columns, text wrapping, borders, alternating row colors, frozen rows)
2.0.5:
- Aligned folder creation with automatic setup in execute_all_scripts
- Removed master HTML references
"""

def setup_xlsx_conversion_logger(debug_dir: str, config: Dict) -> logging.Logger:
    """Set up logger for XLSX conversions."""
    os.makedirs(debug_dir, exist_ok=True)
    log_file = os.path.join(debug_dir, "xlsx_conversion_log.txt")
    logger = logging.getLogger("xlsx_conversion")
    logger.setLevel(logging.INFO)
    handler = logging.FileHandler(log_file, encoding='utf-8')
    handler.setFormatter(logging.Formatter(config['logging']['format']))
    logger.addHandler(handler)
    return logger

def ensure_dependencies(logger: logging.Logger, output_formats: List[str], parallel_processing: bool) -> None:
    """Ensure required dependencies are installed."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl is not installed. Install it using 'pip install openpyxl==3.1.3'")
        raise
    if 'html' in output_formats:
        try:
            import html
        except ImportError:
            logger.error("html module is not available")
            raise

def write_xlsx(xlsx_file: str, fields: List[str], items: List[Dict], config: Dict, logger: logging.Logger) -> None:
    """Write items to an XLSX file with enhanced formatting."""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit Data"
        
        # Define styles
        bold_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        light_gray_fill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')
        
        # Write headers
        for col_idx, field in enumerate(fields, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = field
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = wrap_alignment
        
        # Write data rows
        for row_idx, item in enumerate(items, start=2):
            for col_idx, field in enumerate(fields, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = item.get(field, "")
                cell.border = thin_border
                cell.alignment = wrap_alignment
                # Apply alternating row colors
                if row_idx % 2 == 0:
                    cell.fill = light_gray_fill
        
        # Auto-adjust column widths
        for col_idx, field in enumerate(fields, start=1):
            max_length = max(
                len(str(field)),
                max((len(str(item.get(field, ""))) for item in items), default=0)
            )
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Freeze top row
        ws.freeze_panes = ws['A2']
        
        # Save workbook
        wb.save(xlsx_file)
        logger.info(f"Generated XLSX file: {xlsx_file} with enhanced formatting")
    except Exception as e:
        logger.error(f"Failed to write XLSX file {xlsx_file}: {str(e)}")
        raise

def write_html(html_file: str, fields: List[str], items: List[Dict], config: Dict, logger: logging.Logger) -> None:
    """Write items to an HTML file."""
    try:
        with open(html_file, 'w', encoding='utf-8') as f:
            f.write("<!DOCTYPE html>\n<html>\n<head>\n")
            f.write("<title>Audit Data</title>\n")
            f.write("<style>\n")
            f.write("table { border-collapse: collapse; width: 100%; }\n")
            f.write("th, td { border: 1px solid black; padding: 8px; text-align: left; }\n")
            f.write("th { background-color: #f2f2f2; }\n")
            f.write("</style>\n")
            f.write("</head>\n<body>\n")
            f.write("<h1>Audit Data</h1>\n")
            f.write("<table>\n")
            f.write("<tr>\n")
            for field in fields:
                f.write(f"<th>{escape(field)}</th>\n")
            f.write("</tr>\n")
            for item in items:
                f.write("<tr>\n")
                for field in fields:
                    value = item.get(field, "")
                    f.write(f"<td>{escape(str(value))}</td>\n")
                f.write("</tr>\n")
            f.write("</table>\n")
            f.write("</body>\n</html>")
        logger.info(f"Generated HTML file: {html_file}")
    except Exception as e:
        logger.error(f"Failed to write HTML file {html_file}: {str(e)}")
        raise

def process_file(json_file: str, output_dir: str, config: Dict, logger: logging.Logger) -> None:
    """Process a JSON file to generate XLSX and HTML files."""
    try:
        # Initialize XLSX conversion logger
        xlsx_logger = setup_xlsx_conversion_logger(config['directories']['debug_dir'], config)
        
        # Validate input
        if not os.path.isfile(json_file):
            logger.error(f"JSON file {json_file} does not exist")
            xlsx_logger.error(f"JSON file {json_file} does not exist")
            return
        
        # Get platform-specific output directories
        xlsx_dir = config['output'].get('xlsx_dir', os.path.join(output_dir, "xlsx"))
        html_dir = config['output'].get('html_dir', os.path.join(output_dir, "html"))
        os.makedirs(xlsx_dir, exist_ok=True)
        os.makedirs(html_dir, exist_ok=True)
        
        # Ensure dependencies
        output_formats = config.get('output', {}).get('output_format', ['xlsx'])
        ensure_dependencies(logger, output_formats, config.get('output', {}).get('parallel_processing', True))
        
        # Load JSON data
        start_time = time.time()
        with open(json_file, 'r', encoding='utf-8') as f:
            items = json.load(f)
        load_time = time.time() - start_time
        
        if not items:
            logger.warning(f"No items found in {json_file}")
            xlsx_logger.warning(f"No items found in {json_file}")
            return
        
        # Determine fields, ensuring 'description' is first
        fields = config.get('fields', [])
        if not fields:
            fields = ['description'] + sorted(set(k for item in items for k in item.keys() if k != 'description'))
            logger.info(f"Using detected fields for {json_file}: {fields}")
            xlsx_logger.info(f"Using detected fields for {json_file}: {fields}")
        
        # Generate XLSX
        filename = os.path.basename(json_file).replace("extracted_data_", "").replace(".json", "")
        xlsx_file = os.path.join(xlsx_dir, f"extracted_{filename}.xlsx")
        
        start_write_time = time.time()
        write_xlsx(xlsx_file, fields, items, config, logger)
        write_time = time.time() - start_write_time
        
        # Generate HTML
        if 'html' in output_formats:
            html_file = os.path.join(html_dir, f"extracted_{filename}.html")
            write_html(html_file, fields, items, config, logger)
        
        # Log conversion details
        logger.info(f"Generated XLSX file: {xlsx_file} with {len(items)} rows")
        xlsx_logger.info(f"Converted {json_file} to {xlsx_file}:")
        xlsx_logger.info(f"  Items: {len(items)}")
        xlsx_logger.info(f"  Fields: {len(fields)} ({', '.join(fields)})")
        xlsx_logger.info(f"  JSON load time: {load_time:.3f} seconds")
        xlsx_logger.info(f"  XLSX write time: {write_time:.3f} seconds")
        xlsx_logger.info(f"  Total time: {load_time + write_time:.3f} seconds")
        xlsx_logger.info(f"  Average time per row: {(load_time + write_time) / len(items):.6f} seconds")
        
    except Exception as e:
        logger.error(f"Failed to process {json_file}: {str(e)}")
        xlsx_logger.error(f"Failed to process {json_file}: {str(e)}")
        raise

def main() -> None:
    """Main function for testing audit_extract_helper."""
    parser = argparse.ArgumentParser(description="Process JSON files to XLSX and HTML.")
    parser.add_argument("json_file", help="Input JSON file")
    parser.add_argument("output_dir", help="Output directory for XLSX and HTML")
    args = parser.parse_args()

    try:
        config = {
            "directories": {"debug_dir": "debug"},
            "fields": ["description", "type", "cmd", "value"],
            "output": {"output_format": ["xlsx", "html"], "parallel_processing": True},
            "logging": {
                "level": "INFO",
                "format": "%(asctime)s - %(levelname)s - %(message)s"
            }
        }
        logger = logging.getLogger("audit_extract_helper")
        logger.setLevel(logging.INFO)
        handler = logging.FileHandler(os.path.join(config['directories']['debug_dir'], "extract_helper_main_log.txt"))
        handler.setFormatter(logging.Formatter(config['logging']['format']))
        logger.addHandler(handler)
        
        process_file(args.json_file, args.output_dir, config, logger)
        
    except Exception as e:
        logger.error(f"Execution failed: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()